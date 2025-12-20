"""Question import service for parsing Word and Excel files."""

import io
import re
from typing import Literal

from pydantic import BaseModel


# ============================================
# TYPE DEFINITIONS
# ============================================

FieldType = Literal[
    "text",
    "textarea",
    "email",
    "number",
    "date",
    "select",
    "radio",
    "checkbox",
    "gps",
    "file",
    "phone",
    "url",
    "color",
    "range",
    "rating",
    "signature",
]


class FieldOption(BaseModel):
    """Option for choice fields."""

    label: str
    value: str


class ConditionalRule(BaseModel):
    """Conditional display rule."""

    depends_on: str  # Question number or field ID
    condition: Literal["equals", "not_equals", "contains"]
    value: str


class ParsedQuestion(BaseModel):
    """A parsed question from a document."""

    question_number: int | None = None
    label: str
    type: FieldType
    options: list[FieldOption] | None = None
    required: bool = False
    help_text: str | None = None
    allow_other: bool = False
    section: str | None = None
    conditional: ConditionalRule | None = None


class ImportResult(BaseModel):
    """Result of parsing a document."""

    questions: list[ParsedQuestion]
    sections: list[str]
    warnings: list[str]


# ============================================
# WORD DOCUMENT PARSER
# ============================================


def parse_word_document(file_content: bytes) -> ImportResult:
    """
    Parse a Word document and extract questions.

    Args:
        file_content: Raw bytes of the .docx file

    Returns:
        ImportResult with parsed questions
    """
    from docx import Document

    warnings: list[str] = []
    questions: list[ParsedQuestion] = []
    sections: list[str] = []

    try:
        doc = Document(io.BytesIO(file_content))
    except Exception as e:
        return ImportResult(
            questions=[],
            sections=[],
            warnings=[f"Failed to parse Word document: {e!s}"],
        )

    # Extract all text from paragraphs
    full_text = "\n".join([para.text for para in doc.paragraphs])
    lines = [line.strip() for line in full_text.split("\n") if line.strip()]

    current_section = ""
    i = 0

    while i < len(lines):
        line = lines[i]

        # Check for section headers (e.g., "SECTION A: RESPONDENT BACKGROUND")
        section_match = re.match(r"^SECTION\s+([A-Z]):\s*(.+)", line, re.IGNORECASE)
        if section_match:
            current_section = f"{section_match.group(1)}: {section_match.group(2)}"
            if current_section not in sections:
                sections.append(current_section)
            i += 1
            continue

        # Check for question numbers (e.g., "1.District/Constituency:" or "1. District:")
        question_match = re.match(r"^(\d+)\.?\s*(.+)", line)
        if question_match:
            question_number = int(question_match.group(1))
            question_text = question_match.group(2)

            # Collect full question text and options
            question_lines: list[str] = [question_text]
            option_lines: list[str] = []
            j = i + 1

            # Look ahead for options or continuation
            while j < len(lines):
                next_line = lines[j]

                # If next line is a new question number or section, stop
                if re.match(r"^(\d+)\.?\s", next_line) or re.match(
                    r"^SECTION\s+[A-Z]:", next_line, re.IGNORECASE
                ):
                    break

                # Check if line contains checkbox options
                if "\u2610" in next_line:  # ☐ character
                    option_lines.append(next_line)
                elif len(option_lines) == 0 and not next_line.startswith("If "):
                    # Continuation of question text
                    question_lines.append(next_line)
                elif next_line.startswith("If "):
                    # Conditional follow-up - note it
                    warnings.append(
                        f"Question {question_number}: Conditional follow-up "
                        f'"{next_line}" detected but not fully parsed'
                    )
                j += 1

            # Combine question text and parse
            full_question = " ".join(question_lines).strip()

            # Parse the question
            parsed = _parse_question_line(full_question, option_lines, question_number)
            if parsed:
                parsed.section = current_section if current_section else None
                questions.append(parsed)

            i = j
            continue

        i += 1

    if len(questions) == 0:
        warnings.append(
            "No questions found. Make sure questions are numbered (e.g., 1. Question text)"
        )

    return ImportResult(questions=questions, sections=sections, warnings=warnings)


def _parse_question_line(
    question_text: str,
    option_lines: list[str],
    question_number: int,
) -> ParsedQuestion | None:
    """Parse a single question with its options."""
    # Clean up question text - remove trailing underscores and colons
    label = re.sub(r"[:_]+\s*$", "", question_text)
    label = re.sub(r"\s*_{2,}\s*", " ", label).strip()

    # Check for text field indicators (underscores in the question)
    has_underscores = bool(re.search(r"_{3,}", question_text))

    # Extract options from option lines
    all_options: list[str] = []
    is_multi_select = False

    for opt_line in option_lines:
        # Check for multi-select indicator
        if "tick all that apply" in opt_line.lower() or "select all that apply" in opt_line.lower():
            is_multi_select = True

        # Extract individual options (preceded by ☐)
        option_matches = opt_line.split("\u2610")
        for opt in option_matches:
            cleaned = " ".join(opt.split()).strip()
            if cleaned and "tick all" not in cleaned.lower():
                all_options.append(cleaned)

    # Also check for inline options in the question text itself
    if "\u2610" in question_text:
        inline_options = question_text.split("\u2610")[1:]
        for opt in inline_options:
            cleaned = " ".join(opt.split()).strip()
            if cleaned and cleaned not in all_options:
                all_options.append(cleaned)
        # Remove options from label
        label = re.sub(r"[:_]+\s*$", "", question_text.split("\u2610")[0]).strip()

    # Determine field type
    field_type: FieldType = "text"
    allow_other = False

    if all_options:
        # Check if options indicate multi-select
        options_on_separate_lines = len(option_lines) > 1 or (
            len(option_lines) == 1 and option_lines[0].count("\u2610") > 3
        )

        if is_multi_select or options_on_separate_lines:
            field_type = "checkbox"
        else:
            field_type = "radio"

        # Check for "Other" option with text field
        for idx, opt in enumerate(all_options):
            if "other" in opt.lower() and ("specify" in opt.lower() or "_" in opt):
                allow_other = True
                all_options.pop(idx)
                break
    elif has_underscores:
        field_type = "text"

    # Detect special field types based on question content
    label_lower = label.lower()
    if "email" in label_lower:
        field_type = "email"
    elif "phone" in label_lower or "contact number" in label_lower:
        field_type = "phone"
    elif "date" in label_lower and not all_options:
        field_type = "date"

    # Create options array
    options: list[FieldOption] = []
    for idx, opt in enumerate(all_options):
        value = re.sub(r"[^a-z0-9]+", "_", opt.lower())
        value = value.strip("_") or f"option_{idx + 1}"
        options.append(FieldOption(label=opt, value=value))

    return ParsedQuestion(
        question_number=question_number,
        label=label,
        type=field_type,
        options=options if options else None,
        required=False,  # Default to false, user can change in form builder
        allow_other=allow_other,
    )


# ============================================
# EXCEL FILE PARSER
# ============================================


def parse_excel_file(file_content: bytes) -> ImportResult:
    """
    Parse an Excel file and extract questions.

    Args:
        file_content: Raw bytes of the .xlsx file

    Returns:
        ImportResult with parsed questions
    """
    from openpyxl import load_workbook

    warnings: list[str] = []
    questions: list[ParsedQuestion] = []
    sections: list[str] = []

    try:
        wb = load_workbook(io.BytesIO(file_content), read_only=True, data_only=True)
    except Exception as e:
        return ImportResult(
            questions=[],
            sections=[],
            warnings=[f"Failed to parse Excel file: {e!s}"],
        )

    # Get the first sheet
    if not wb.sheetnames:
        return ImportResult(
            questions=[],
            sections=[],
            warnings=["No sheets found in the Excel file"],
        )

    ws = wb.active
    if ws is None:
        return ImportResult(
            questions=[],
            sections=[],
            warnings=["No active sheet in the Excel file"],
        )

    # Read headers from first row
    headers: list[str] = []
    for cell in ws[1]:
        headers.append(str(cell.value or "").lower().strip())

    # Check for required columns
    if "question" not in headers:
        warnings.append("Missing required column: 'question'. Please use the template.")
    if "type" not in headers:
        warnings.append("Missing 'type' column. Field types will be inferred.")

    # Get column indices
    col_indices = {header: idx for idx, header in enumerate(headers)}

    # Parse each row (skip header)
    row_num = 2
    for row in ws.iter_rows(min_row=2):
        try:
            row_values = [cell.value for cell in row]

            # Get question text
            question_idx = col_indices.get("question")
            if question_idx is None or not row_values[question_idx]:
                row_num += 1
                continue

            question_text = str(row_values[question_idx]).strip()

            # Get type
            type_idx = col_indices.get("type")
            field_type = _normalize_field_type(
                str(row_values[type_idx]) if type_idx is not None and row_values[type_idx] else "text"
            )

            # Get options
            options: list[FieldOption] | None = None
            options_idx = col_indices.get("options")
            if options_idx is not None and row_values[options_idx]:
                option_strings = str(row_values[options_idx]).split("|")
                options = []
                for idx, opt in enumerate(option_strings):
                    opt = opt.strip()
                    if opt:
                        value = re.sub(r"[^a-z0-9]+", "_", opt.lower()).strip("_") or f"option_{idx + 1}"
                        options.append(FieldOption(label=opt, value=value))

                # If options provided but no type specified, default to radio
                if options and field_type == "text":
                    field_type = "radio"

            # Get other fields
            required = _parse_boolean(
                row_values[col_indices["required"]]
                if "required" in col_indices and col_indices["required"] < len(row_values)
                else None
            )
            help_text = (
                str(row_values[col_indices["help_text"]]).strip()
                if "help_text" in col_indices
                and col_indices["help_text"] < len(row_values)
                and row_values[col_indices["help_text"]]
                else None
            )
            allow_other = _parse_boolean(
                row_values[col_indices["allow_other"]]
                if "allow_other" in col_indices and col_indices["allow_other"] < len(row_values)
                else None
            )
            section = (
                str(row_values[col_indices["section"]]).strip()
                if "section" in col_indices
                and col_indices["section"] < len(row_values)
                and row_values[col_indices["section"]]
                else None
            )
            question_number = (
                int(row_values[col_indices["question_number"]])
                if "question_number" in col_indices
                and col_indices["question_number"] < len(row_values)
                and row_values[col_indices["question_number"]]
                else None
            )

            # Parse conditional
            conditional: ConditionalRule | None = None
            depends_on_idx = col_indices.get("depends_on")
            show_when_idx = col_indices.get("show_when")
            if (
                depends_on_idx is not None
                and show_when_idx is not None
                and depends_on_idx < len(row_values)
                and show_when_idx < len(row_values)
                and row_values[depends_on_idx]
                and row_values[show_when_idx]
            ):
                conditional = ConditionalRule(
                    depends_on=str(row_values[depends_on_idx]).strip(),
                    condition="equals",
                    value=str(row_values[show_when_idx]).strip(),
                )

            # Add to sections list
            if section and section not in sections:
                sections.append(section)

            questions.append(
                ParsedQuestion(
                    question_number=question_number,
                    label=question_text,
                    type=field_type,
                    options=options if options else None,
                    required=required,
                    help_text=help_text,
                    allow_other=allow_other,
                    section=section,
                    conditional=conditional,
                )
            )

        except Exception as e:
            warnings.append(f"Row {row_num}: {e!s}")

        row_num += 1

    wb.close()

    if not questions:
        warnings.append("No questions found in the Excel file.")

    return ImportResult(questions=questions, sections=sections, warnings=warnings)


def _normalize_field_type(type_str: str) -> FieldType:
    """Normalize field type string to valid FieldType."""
    type_lower = type_str.lower().strip()

    type_map: dict[str, FieldType] = {
        "text": "text",
        "short text": "text",
        "short answer": "text",
        "textarea": "textarea",
        "long text": "textarea",
        "long answer": "textarea",
        "paragraph": "textarea",
        "email": "email",
        "email address": "email",
        "number": "number",
        "numeric": "number",
        "integer": "number",
        "date": "date",
        "select": "select",
        "dropdown": "select",
        "radio": "radio",
        "radio button": "radio",
        "single choice": "radio",
        "single select": "radio",
        "checkbox": "checkbox",
        "checkboxes": "checkbox",
        "multiple choice": "checkbox",
        "multi select": "checkbox",
        "gps": "gps",
        "location": "gps",
        "coordinates": "gps",
        "file": "file",
        "upload": "file",
        "attachment": "file",
        "phone": "phone",
        "phone number": "phone",
        "telephone": "phone",
        "url": "url",
        "website": "url",
        "link": "url",
        "color": "color",
        "color picker": "color",
        "range": "range",
        "slider": "range",
        "rating": "rating",
        "stars": "rating",
        "signature": "signature",
        "sign": "signature",
    }

    return type_map.get(type_lower, "text")


def _parse_boolean(value: str | bool | int | None) -> bool:
    """Parse boolean from various formats."""
    if value is None:
        return False
    if isinstance(value, bool):
        return value
    if isinstance(value, int):
        return value == 1
    val_str = str(value).lower().strip()
    return val_str in ("true", "yes", "1", "y")


# ============================================
# FILE TYPE DETECTION
# ============================================


def detect_file_type(filename: str, content_type: str | None = None) -> str | None:
    """
    Detect file type from filename and content type.

    Returns:
        'word' for Word documents, 'excel' for Excel files, None if unsupported
    """
    extension = filename.lower().split(".")[-1] if "." in filename else ""

    if extension in ("docx",):
        return "word"
    if extension in ("xlsx", "xls"):
        return "excel"

    # Fallback to content type
    if content_type:
        if "wordprocessingml" in content_type or "msword" in content_type:
            return "word"
        if "spreadsheetml" in content_type or "ms-excel" in content_type:
            return "excel"

    return None


# ============================================
# MAIN PARSE FUNCTION
# ============================================


def parse_question_file(
    file_content: bytes,
    filename: str,
    content_type: str | None = None,
) -> ImportResult:
    """
    Parse a file and extract questions.

    Args:
        file_content: Raw bytes of the file
        filename: Original filename (for type detection)
        content_type: MIME type (optional, for type detection)

    Returns:
        ImportResult with parsed questions
    """
    file_type = detect_file_type(filename, content_type)

    if file_type == "word":
        return parse_word_document(file_content)
    if file_type == "excel":
        return parse_excel_file(file_content)

    return ImportResult(
        questions=[],
        sections=[],
        warnings=[f"Unsupported file type: {filename}. Please upload a .docx or .xlsx file."],
    )
