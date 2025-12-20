"""AI-powered question refinement service using OpenAI GPT-4o."""

import json
import logging
from typing import Literal

from openai import AsyncOpenAI
from pydantic import BaseModel

from app.core.config import settings

logger = logging.getLogger(__name__)


# ============================================
# TYPE DEFINITIONS
# ============================================

FieldType = Literal[
    "text",
    "textarea",
    "email",
    "number",
    "date",
    "select",
    "radio",
    "checkbox",
    "gps",
    "file",
    "phone",
    "url",
    "color",
    "range",
    "rating",
    "signature",
]


class FieldOption(BaseModel):
    """Option for choice fields."""

    label: str
    value: str


class ConditionalRule(BaseModel):
    """Conditional display rule."""

    depends_on: str
    condition: Literal["equals", "not_equals", "contains"]
    value: str


class RefinedQuestion(BaseModel):
    """A refined question from AI processing."""

    question_number: int | None = None
    label: str
    type: FieldType
    options: list[FieldOption] | None = None
    required: bool = False
    help_text: str | None = None
    allow_other: bool = False
    section: str | None = None
    conditional: ConditionalRule | None = None


class AIRefineResult(BaseModel):
    """Result of AI refinement."""

    success: bool
    questions: list[RefinedQuestion]
    sections: list[str]
    summary: str | None = None
    error: str | None = None


# ============================================
# SYSTEM PROMPT
# ============================================

SYSTEM_PROMPT = """You are an expert form builder assistant. Your task is to analyze raw document text containing survey questions and extract them into a structured format suitable for a form builder application.

## Field Types Available

You must classify each question into one of these field types:
- **text**: Short single-line text input (names, short answers)
- **textarea**: Long multi-line text input (comments, descriptions, explanations)
- **email**: Email address input
- **phone**: Phone number input
- **number**: Numeric input (age as number, quantities, scores)
- **date**: Date picker
- **radio**: Single choice from options (Yes/No, rating scales, mutually exclusive options)
- **checkbox**: Multiple choice - can select many options (interests, services used, "tick all that apply")
- **select**: Dropdown menu (alternative to radio for many options)
- **rating**: Star rating (satisfaction, quality scores 1-5)
- **gps**: GPS location capture
- **file**: File upload
- **signature**: Signature capture

## Classification Rules

1. **Radio vs Checkbox**:
   - Use **radio** when only ONE option can be selected (Yes/No, gender, age groups, Likert scales)
   - Use **checkbox** when MULTIPLE options can be selected ("tick all that apply", services used, interests)

2. **Text vs Textarea**:
   - Use **text** for short answers (names, titles, single-line responses)
   - Use **textarea** for long responses ("explain why", "describe", "comments", "additional feedback")

3. **Required Fields**:
   - Mark as required if the question seems essential or is marked with asterisk/required indicator

4. **Conditional Questions**:
   - Detect "If yes...", "If no...", "If applicable..." patterns
   - Set the conditional.depends_on to the question number it depends on
   - Set conditional.value to the trigger value (e.g., "yes", "no")

5. **Allow Other**:
   - Set allow_other=true when options include "Other (specify)", "Other (please specify)", etc.
   - Remove the "Other" option from the options list when allow_other is true

6. **Sections**:
   - Detect section headers like "SECTION A:", "Part 1:", "Demographics:", etc.
   - Assign questions to their appropriate sections

## Output Format

Return a valid JSON object with this structure:
{
  "questions": [
    {
      "question_number": 1,
      "label": "What is your full name?",
      "type": "text",
      "options": null,
      "required": true,
      "help_text": null,
      "allow_other": false,
      "section": "Personal Information",
      "conditional": null
    },
    {
      "question_number": 2,
      "label": "What is your gender?",
      "type": "radio",
      "options": [
        {"label": "Male", "value": "male"},
        {"label": "Female", "value": "female"}
      ],
      "required": true,
      "help_text": null,
      "allow_other": true,
      "section": "Demographics",
      "conditional": null
    }
  ],
  "sections": ["Personal Information", "Demographics"],
  "summary": "Extracted 15 questions across 3 sections"
}

## Important Notes

- Generate clean, lowercase snake_case values for options (e.g., "Very Satisfied" -> "very_satisfied")
- Clean up question labels - remove numbering, trailing colons, underscores for blanks
- Detect Likert scales (Strongly Agree to Strongly Disagree) and use radio type
- For age ranges like "18-25, 26-35, 36-45, 46+", use radio (single choice)
- If you see ☐ symbols, these are checkbox/radio indicators
- Return ONLY the JSON object, no markdown code blocks or explanations"""


# ============================================
# AI REFINEMENT SERVICE
# ============================================


async def refine_questions_with_ai(
    document_text: str,
    existing_questions: list[dict] | None = None,
) -> AIRefineResult:
    """
    Use GPT-4o to extract and refine questions from document text.

    Args:
        document_text: Raw text content from the document
        existing_questions: Optional list of already-parsed questions for context

    Returns:
        AIRefineResult with refined questions
    """
    if not settings.OPENAI_API_KEY:
        return AIRefineResult(
            success=False,
            questions=[],
            sections=[],
            error="OpenAI API key not configured. Please set OPENAI_API_KEY environment variable.",
        )

    client = AsyncOpenAI(api_key=settings.OPENAI_API_KEY)

    # Build user prompt
    user_prompt = f"""Please analyze the following document and extract all questions into a structured format.

## Document Content:

{document_text}

---

Please extract all questions and return them as a JSON object following the format specified in your instructions."""

    # Add context about existing parsing if available
    if existing_questions:
        user_prompt += f"""

## Previous Parsing Attempt (for reference):

The document was previously parsed and {len(existing_questions)} questions were found. Please improve upon this parsing, ensuring correct field types and structure.

Previous questions found:
{json.dumps(existing_questions[:5], indent=2)}
{"..." if len(existing_questions) > 5 else ""}"""

    try:
        response = await client.chat.completions.create(
            model=settings.OPENAI_MODEL,
            messages=[
                {"role": "system", "content": SYSTEM_PROMPT},
                {"role": "user", "content": user_prompt},
            ],
            temperature=0.1,  # Low temperature for consistent, accurate extraction
            max_tokens=4096,
            response_format={"type": "json_object"},
        )

        # Parse response
        content = response.choices[0].message.content
        if not content:
            return AIRefineResult(
                success=False,
                questions=[],
                sections=[],
                error="Empty response from AI model",
            )

        # Parse JSON response
        try:
            data = json.loads(content)
        except json.JSONDecodeError as e:
            logger.error(f"Failed to parse AI response as JSON: {e}")
            return AIRefineResult(
                success=False,
                questions=[],
                sections=[],
                error=f"Failed to parse AI response: {e}",
            )

        # Convert to RefinedQuestion objects
        questions: list[RefinedQuestion] = []
        for q in data.get("questions", []):
            try:
                # Parse options
                options = None
                if q.get("options"):
                    options = [
                        FieldOption(label=o["label"], value=o["value"])
                        for o in q["options"]
                    ]

                # Parse conditional
                conditional = None
                if q.get("conditional"):
                    conditional = ConditionalRule(
                        depends_on=str(q["conditional"]["depends_on"]),
                        condition=q["conditional"].get("condition", "equals"),
                        value=str(q["conditional"]["value"]),
                    )

                questions.append(
                    RefinedQuestion(
                        question_number=q.get("question_number"),
                        label=q["label"],
                        type=q.get("type", "text"),
                        options=options,
                        required=q.get("required", False),
                        help_text=q.get("help_text"),
                        allow_other=q.get("allow_other", False),
                        section=q.get("section"),
                        conditional=conditional,
                    )
                )
            except Exception as e:
                logger.warning(f"Failed to parse question: {e}, skipping")
                continue

        sections = data.get("sections", [])
        summary = data.get("summary")

        return AIRefineResult(
            success=True,
            questions=questions,
            sections=sections,
            summary=summary,
        )

    except Exception as e:
        logger.error(f"OpenAI API error: {e}")
        return AIRefineResult(
            success=False,
            questions=[],
            sections=[],
            error=f"AI processing error: {e!s}",
        )


async def extract_text_for_ai(file_content: bytes, file_type: str) -> str:
    """
    Extract raw text from document for AI processing.

    Args:
        file_content: Raw bytes of the file
        file_type: 'word' or 'excel'

    Returns:
        Extracted text content
    """
    import io

    if file_type == "word":
        from docx import Document

        doc = Document(io.BytesIO(file_content))
        paragraphs = [para.text for para in doc.paragraphs if para.text.strip()]
        return "\n".join(paragraphs)

    elif file_type == "excel":
        from openpyxl import load_workbook

        wb = load_workbook(io.BytesIO(file_content), read_only=True, data_only=True)
        ws = wb.active
        if not ws:
            return ""

        rows = []
        for row in ws.iter_rows():
            row_values = [str(cell.value or "") for cell in row]
            if any(v.strip() for v in row_values):
                rows.append(" | ".join(row_values))

        wb.close()
        return "\n".join(rows)

    return ""
